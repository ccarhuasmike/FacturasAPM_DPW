from sharepoint_conexion import upload_file_to_sharepoint_v2
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.worksheet.table import Table
from sharepoint_conexion import (
    get_graph_token,
    get_drive_item_id_from_url
)
import requests
import logging

log_filename = os.path.join(os.path.dirname(__file__), 'procesamiento_log.txt')
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(log_filename, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
def descargar_excel_desde_sharepoint(shared_file_url, destino_local):
    token = get_graph_token()
    item_data = get_drive_item_id_from_url(shared_file_url, token)

    drive_id = item_data['parentReference']['driveId']
    item_id = item_data['id']
    nombre_archivo = item_data['name']

    url_download = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content"
    headers = {"Authorization": f"Bearer {token}"}

    logging.info(f"Descargando archivo '{nombre_archivo}' desde SharePoint...")
    response = requests.get(url_download, headers=headers)
    response.raise_for_status()

    ruta_guardado = destino_local if destino_local.endswith(".xlsx") else f"{destino_local}\\{nombre_archivo}"

    with open(ruta_guardado, "wb") as f:
        f.write(response.content)

    logging.info(f"Archivo descargado exitosamente a: {ruta_guardado}")


def procesar_excel_en_sharepoint_y_limpiar_local_temporal(
    shared_file_url,
    archivo_local_path,
    hoja_objetivo
):

    logging.info("Descargando archivo desde SharePoint...")

    token = get_graph_token()
    item_data = get_drive_item_id_from_url(shared_file_url, token)

    drive_id = item_data['parentReference']['driveId']
    item_id = item_data['id']
    file_name = item_data['name']

    url_download = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content"
    headers = {"Authorization": f"Bearer {token}"}
    response = requests.get(url_download, headers=headers)
    response.raise_for_status()

    local_temp_path = f"temp_{file_name}"
    with open(local_temp_path, "wb") as f:
        f.write(response.content)

    logging.info("Cargando archivos...")

    df_local = pd.read_excel(archivo_local_path, sheet_name=hoja_objetivo, dtype=str)
    logging.info(f"Filas originales leídas desde archivo local: {len(df_local)}")

    # -------------------------------
    # LIMPIEZA DE FILAS VACÍAS
    # -------------------------------
    df_local.dropna(how="all", inplace=True)
    df_local = df_local[
        ~df_local.apply(lambda row: row.astype(str).str.strip().eq("").all(), axis=1)
    ]
    df_local.reset_index(drop=True, inplace=True)
    logging.info(f"Filas válidas para insertar después de limpieza: {len(df_local)}")

    if df_local.empty:
        logging.warning("No hay registros válidos en el archivo local, se omite la carga.")
        return

    wb = load_workbook(local_temp_path)
    ws = wb[hoja_objetivo]

    if not ws.tables:
        logging.error(f"No se encontró ninguna tabla en la hoja '{hoja_objetivo}'")
        return

    tabla_name = list(ws.tables.keys())[0]
    if len(ws.tables) > 1:
        logging.warning(f"Hay más de una tabla, se usará la primera: '{tabla_name}'")

    tabla = ws.tables[tabla_name]
    min_col, start_row, max_col, end_row = range_boundaries(tabla.ref)
    logging.info(f"Rango actual de la tabla: {tabla.ref}")

    # -------------------------------
    # DETECTAR ÚLTIMA FILA REAL DENTRO DE LA TABLA
    # -------------------------------
    ultima_fila_real = start_row
    for fila in range(end_row, start_row, -1):
        valores = [ws.cell(row=fila, column=c).value for c in range(min_col, max_col + 1)]
        if any(v not in (None, "") for v in valores):
            ultima_fila_real = fila
            break

    logging.info(f"Última fila REAL con datos dentro de la tabla: {ultima_fila_real}")

    # -------------------------------
    # INSERTAR REGISTROS
    # -------------------------------
    fila_actual = ultima_fila_real + 1
    logging.info(f"Insertando registros desde la fila: {fila_actual}")

    registros_insertados = 0
    for _, row in df_local.iterrows():
        for col_idx, value in enumerate(row.tolist(), start=min_col):
            ws.cell(row=fila_actual, column=col_idx, value=value)
        fila_actual += 1
        registros_insertados += 1

    logging.info(f"Total de registros insertados: {registros_insertados}")

    # -------------------------------
    # EXPANDIR TABLA
    # -------------------------------
    nueva_ultima_fila = fila_actual - 1
    nuevo_ref = f"{get_column_letter(min_col)}{start_row}:{get_column_letter(max_col)}{nueva_ultima_fila}"
    tabla.ref = nuevo_ref
    logging.info(f"Nuevo rango expandido de la tabla: {nuevo_ref}")

    # -------------------------------
    # GUARDAR Y SUBIR
    # -------------------------------
    logging.info("Guardando archivo temporal modificado...")
    wb.save(local_temp_path)

    logging.info("Subiendo archivo modificado a SharePoint...")
    with open(local_temp_path, "rb") as f:
        upload_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content"
        response = requests.put(upload_url, headers=headers, data=f)
        logging.info(f"Respuesta de subida: {response.status_code} - {response.reason}")
        response.raise_for_status()

    logging.info("Limpiando datos del archivo local...")
    limpiar_excel_local_conservando_formato(archivo_local_path)

    os.remove(local_temp_path)
    logging.info("Proceso completado con éxito.")


def procesar_excel_en_sharepoint_y_limpiar_local(
    shared_file_url,
    archivo_local_path,
    hoja_objetivo
):

    logging.info("Descargando archivo desde SharePoint...")

    token = get_graph_token()
    item_data = get_drive_item_id_from_url(shared_file_url, token)

    drive_id = item_data['parentReference']['driveId']
    item_id = item_data['id']
    file_name = item_data['name']

    url_download = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content"
    headers = {"Authorization": f"Bearer {token}"}
    response = requests.get(url_download, headers=headers)
    response.raise_for_status()

    local_temp_path = f"temp_{file_name}"
    with open(local_temp_path, "wb") as f:
        f.write(response.content)

    logging.info("Cargando archivos...")

    df_local = pd.read_excel(archivo_local_path, sheet_name=hoja_objetivo, dtype=str)
    logging.info(f"Filas originales leídas desde archivo local: {len(df_local)}")

    # -------------------------------
    # LIMPIEZA DE FILAS VACÍAS
    # -------------------------------
    df_local.dropna(how="all", inplace=True)
    df_local = df_local[
        ~df_local.apply(lambda row: row.astype(str).str.strip().eq("").all(), axis=1)
    ]
    df_local.reset_index(drop=True, inplace=True)
    logging.info(f"Filas válidas para insertar después de limpieza: {len(df_local)}")

    if df_local.empty:
        logging.warning("No hay registros válidos en el archivo local, se omite la carga.")
        return

    wb = load_workbook(local_temp_path)
    ws = wb[hoja_objetivo]

    if not ws.tables:
        logging.error(f"No se encontró ninguna tabla en la hoja '{hoja_objetivo}'")
        return

    tabla_name = list(ws.tables.keys())[0]
    if len(ws.tables) > 1:
        logging.warning(f"Hay más de una tabla, se usará la primera: '{tabla_name}'")

    tabla = ws.tables[tabla_name]
    min_col, start_row, max_col, end_row = range_boundaries(tabla.ref)
    logging.info(f"Rango actual de la tabla: {tabla.ref}")

    # -------------------------------
    # DETECTAR ÚLTIMA FILA REAL DENTRO DE LA TABLA
    # -------------------------------
    ultima_fila_real = start_row
    for fila in range(end_row, start_row, -1):
        valores = [ws.cell(row=fila, column=c).value for c in range(min_col, max_col + 1)]
        if any(v not in (None, "") for v in valores):
            ultima_fila_real = fila
            break

    logging.info(f"Última fila REAL con datos dentro de la tabla: {ultima_fila_real}")

    # -------------------------------
    # INSERTAR REGISTROS
    # -------------------------------
    fila_actual = ultima_fila_real + 1
    logging.info(f"Insertando registros desde la fila: {fila_actual}")

    registros_insertados = 0
    for _, row in df_local.iterrows():
        for col_idx, value in enumerate(row.tolist(), start=min_col):
            ws.cell(row=fila_actual, column=col_idx, value=value)
        fila_actual += 1
        registros_insertados += 1

    logging.info(f"Total de registros insertados: {registros_insertados}")

    # -------------------------------
    # EXPANDIR TABLA
    # -------------------------------
    nueva_ultima_fila = fila_actual - 1
    nuevo_ref = f"{get_column_letter(min_col)}{start_row}:{get_column_letter(max_col)}{nueva_ultima_fila}"
    tabla.ref = nuevo_ref
    logging.info(f"Nuevo rango expandido de la tabla: {nuevo_ref}")

    # -------------------------------
    # GUARDAR Y SUBIR
    # -------------------------------
    logging.info("Guardando archivo temporal modificado...")
    wb.save(local_temp_path)

    logging.info("Subiendo archivo modificado a SharePoint...")
    with open(local_temp_path, "rb") as f:
        upload_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{item_id}/content"
        response = requests.put(upload_url, headers=headers, data=f)
        logging.info(f"Respuesta de subida: {response.status_code} - {response.reason}")
        response.raise_for_status()

    logging.info("Limpiando datos del archivo local...")
    limpiar_excel_local_conservando_formato(archivo_local_path)

    os.remove(local_temp_path)
    logging.info("Proceso completado con éxito.")

def limpiar_excel_local_conservando_formato(path_excel):
    wb = load_workbook(path_excel)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(min_row=2):  # Mantiene encabezado (fila 1)
            for cell in row:
                cell.value = ""
    wb.save(path_excel)
    logging.info(f"Archivo local limpiado correctamente: {path_excel}")

def subir_pdfs_en_lote_a_sharepoint(local_folder_path, shared_folder_url, subfolder_path=""):
    archivos_pdf = [f for f in os.listdir(local_folder_path) if f.lower().endswith(".pdf")]

    if not archivos_pdf:
        logging.warning("No se encontraron archivos PDF en la carpeta local.")
        return

    logging.info(f"Subiendo {len(archivos_pdf)} archivos PDF desde: {local_folder_path}")

    for nombre_archivo in archivos_pdf:
        ruta_completa = os.path.join(local_folder_path, nombre_archivo)

        try:
            logging.info(f"Subiendo: {nombre_archivo}...")
            upload_file_to_sharepoint_v2(
                shared_folder_url=shared_folder_url,
                local_file_path=ruta_completa,
                subfolder_path=subfolder_path
            )

            # Eliminar el archivo local solo si se subió sin errores
            os.remove(ruta_completa)
            logging.info(f"Eliminado de local: {nombre_archivo}")

        except Exception as e:
            logging.error(f"Error al subir {nombre_archivo}: {str(e)} — NO eliminado.")

    logging.info("Todos los archivos PDF han sido procesados.")

def descargar_y_eliminar_archivos_sharepoint(shared_folder_url, carpeta_destino_local):
    token = get_graph_token()
    item_data = get_drive_item_id_from_url(shared_folder_url, token)

    drive_id = item_data['parentReference']['driveId']
    folder_id = item_data['id']

    # Crear carpeta local si no existe
    os.makedirs(carpeta_destino_local, exist_ok=True)

    # Obtener lista de archivos dentro del folder
    url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{folder_id}/children"
    headers = {"Authorization": f"Bearer {token}"}
    response = requests.get(url, headers=headers)
    response.raise_for_status()

    archivos = response.json().get("value", [])

    if not archivos:
        logging.warning("No se encontraron archivos en el folder de SharePoint.")
        return

    logging.info(f"Se encontraron {len(archivos)} archivos para descargar y eliminar.")

    for archivo in archivos:
        nombre_archivo = archivo['name']
        archivo_id = archivo['id']
        archivo_local_path = os.path.join(carpeta_destino_local, nombre_archivo)

        # Descargar archivo
        url_descarga = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{archivo_id}/content"
        logging.info(f"Descargando: {nombre_archivo}")
        resp_descarga = requests.get(url_descarga, headers=headers)
        resp_descarga.raise_for_status()

        with open(archivo_local_path, "wb") as f:
            f.write(resp_descarga.content)
        logging.info(f"Guardado en: {archivo_local_path}")

        # Eliminar archivo
        url_delete = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{archivo_id}"
        resp_delete = requests.delete(url_delete, headers=headers)

        if resp_delete.status_code == 204:
            logging.info(f"Eliminado de SharePoint: {nombre_archivo}")
        else:
             logging.error(f"Error al eliminar {nombre_archivo}: {resp_delete.status_code} - {resp_delete.text}")

    logging.info("Proceso de descarga y eliminación completado.")

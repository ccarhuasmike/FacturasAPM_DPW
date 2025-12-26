# procesamiento_v4.py
import os
import re
import PyPDF2
import pandas as pd
from datetime import datetime
import shutil
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
import zipfile
import time
import logging

# ==================== CONFIGURACIÓN DE RUTAS ====================
input_folder = r"D:\Python\ProcesamientoFaturas\Proyecto - Extraccion y Carga de Facturas APM & DPW\Files\APM\Input"
output_apm_pdf = r"D:\Python\ProcesamientoFaturas\Proyecto - Extraccion y Carga de Facturas APM & DPW\Files\APM\Resultado_PDF"
output_apm_excel = r"D:\Python\ProcesamientoFaturas\Proyecto - Extraccion y Carga de Facturas APM & DPW\Files\APM\Recobro Servicios Portuarios APM 2025.xlsx"
base_tipo_servicio_apm = r"D:\Python\ProcesamientoFaturas\Proyecto - Extraccion y Carga de Facturas APM & DPW\Files\APM\Base_tipo_servicio.xlsx"


input_folder_dpw = r"D:\Python\ProcesamientoFaturas\Proyecto - Extraccion y Carga de Facturas APM & DPW\Files\DPW\Input"
output_dpw_pdf = r"D:\Python\ProcesamientoFaturas\Proyecto - Extraccion y Carga de Facturas APM & DPW\Files\DPW\Resultado_PDF"
output_dpw_excel = r"D:\Python\ProcesamientoFaturas\Proyecto - Extraccion y Carga de Facturas APM & DPW\Files\DPW\Recobro servicios portuarios DPW 2025.xlsx"
base_tipo_servicio_dpw = r"D:\Python\ProcesamientoFaturas\Proyecto - Extraccion y Carga de Facturas APM & DPW\Files\DPW\Base_tipo_servicio.xlsx"


# ==================== LOGS ====================
log_filename = os.path.join(os.path.dirname(__file__), 'procesamiento_log.txt')
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler(log_filename, encoding='utf-8'),
        logging.StreamHandler()
    ]
)

# ====================   C O N F I G U R A C I Ó N     R E G U L A R   ====================



def apm_extraer_texto_pdf(ruta_pdf):
    try:
        with open(ruta_pdf, 'rb') as pdf_file:
            reader = PyPDF2.PdfReader(pdf_file)
            return "".join(page.extract_text() for page in reader.pages)
    except Exception as e:
        logging.error(f"Error al leer el archivo PDF: {e}")
        return None

def apm_extraer_cod_factura(texto_pdf):
    patron = r"N°\s*([A-Za-z0-9-]+)\s*FACTURA ELECTRONICA"
    coincidencia = re.search(patron, texto_pdf)
    return coincidencia.group(1) if coincidencia else None

def apm_extraer_fecha(texto_pdf):
    patron = r"F\. De Emisión\s*:\s*(\d{4}-\d{2}-\d{2})"
    coincidencia = re.search(patron, texto_pdf)
    if coincidencia:
        try:
            return datetime.strptime(coincidencia.group(1), "%Y-%m-%d").strftime("%d-%m-%Y")
        except ValueError:
            logging.error(f"Formato de fecha no reconocido: {coincidencia.group(1)}")
    return None

def apm_procesar_pdfs():
    os.makedirs(output_apm_pdf, exist_ok=True)
    
    archivos_pdf = [f for f in os.listdir(input_folder) if f.lower().endswith('.pdf')]
    archivos_excel = [f for f in os.listdir(input_folder) if f.lower().endswith('.xlsx')]
    
    if not archivos_pdf or not archivos_excel:
        logging.warning("No se encontraron ambos archivos (PDF y Excel) en la carpeta.")
        return False
    
    pdf_original = archivos_pdf[0]
    ruta_pdf_original = os.path.join(input_folder, pdf_original)
    
    nombre_base = os.path.splitext(pdf_original)[0]
    cod_factura = nombre_base.split('_')[0]
    logging.info(f"Código de factura extraído: {cod_factura}")
    
    nuevo_nombre_pdf = f"{cod_factura}.pdf"
    ruta_pdf_nuevo = os.path.join(input_folder, nuevo_nombre_pdf)
    
    try:
        os.rename(ruta_pdf_original, ruta_pdf_nuevo)
        logging.info(f"Archivo PDF renombrado a: {nuevo_nombre_pdf}")
    except Exception as e:
        logging.error(f"Error al renombrar el archivo PDF: {e}")
        return False
    
    texto_pdf = apm_extraer_texto_pdf(ruta_pdf_nuevo)
    fecha_emision = apm_extraer_fecha(texto_pdf) if texto_pdf else None
    
    if not fecha_emision:
        logging.error("No se pudo determinar la fecha de emisión.")
        return False
    
    excel_original = archivos_excel[0]
    ruta_excel_original = os.path.join(input_folder, excel_original)
    
    nombre_base_excel = os.path.splitext(excel_original)[0]
    nuevo_nombre_excel = f"{nombre_base_excel}_{cod_factura}_{fecha_emision}.xlsx"
    ruta_excel_nuevo = os.path.join(input_folder, nuevo_nombre_excel)
    
    try:
        os.rename(ruta_excel_original, ruta_excel_nuevo)
        logging.info(f"Archivo Excel renombrado a: {nuevo_nombre_excel}")
    except Exception as e:
        logging.error(f"Error al renombrar el archivo Excel: {e}")
    
    try:
        shutil.move(ruta_pdf_nuevo, os.path.join(output_apm_pdf, nuevo_nombre_pdf))
        logging.info(f"Archivo PDF movido a: {output_apm_pdf}")
        return True
    except Exception as e:
        logging.error(f"Error al mover el archivo PDF: {e}")
        return False

def apm_procesar_excels():
    try:
        wb = load_workbook(output_apm_excel)
        
        if 'APMT' not in wb.sheetnames:
            raise ValueError("La hoja 'APMT' no existe en el archivo base.")
        
        ws = wb['APMT']
        
        next_row = 1
        while ws.cell(row=next_row, column=1).value is not None:
            next_row += 1

        style_reference = []
        for cell in ws[2]:
            style_reference.append({
                'fill': PatternFill(start_color=cell.fill.start_color, end_color=cell.fill.end_color, fill_type=cell.fill.fill_type) if cell.fill else None,
                'font': Font(name=cell.font.name, size=cell.font.size, bold=cell.font.bold, italic=cell.font.italic, color=cell.font.color) if cell.font else None,
                'border': Border(
                    left=Side(border_style=cell.border.left.border_style, color=cell.border.left.color) if cell.border.left else None,
                    right=Side(border_style=cell.border.right.border_style, color=cell.border.right.color) if cell.border.right else None,
                    top=Side(border_style=cell.border.top.border_style, color=cell.border.top.color) if cell.border.top else None,
                    bottom=Side(border_style=cell.border.bottom.border_style, color=cell.border.bottom.color) if cell.border.bottom else None
                ) if cell.border else None,
                'alignment': Alignment(horizontal=cell.alignment.horizontal, vertical=cell.alignment.vertical, wrap_text=cell.alignment.wrap_text) if cell.alignment else None,
                'number_format': cell.number_format
            })

        for filename in os.listdir(input_folder):
            if filename.endswith(".xlsx"):
                filepath = os.path.join(input_folder, filename)
                df_input = pd.read_excel(filepath, header=8)

                df_tipo_servicio = pd.read_excel(base_tipo_servicio_apm, dtype=str)

                df_input['Tarifa'] = df_input['Tarifa'].astype(str).str.strip()

                df_tipo_servicio['Tarifa'] = df_tipo_servicio['Tarifa'].astype(str).str.strip()

                df_input['BL'] = df_input['BL'].astype(str).replace('nan', '').str.strip() if 'BL' in df_input.columns else ""

                df_input.insert(0, "Proceso", df_input['BL'].apply(lambda x: "EMBARQUE" if x == "" else "DESCARGA"))

                partes = filename.replace(".xlsx", "").split("_")
                invoice = partes[-2]
                fecha_factura = datetime.strptime(partes[-1], "%d-%m-%Y").strftime("%d/%m/%Y")

                df_input.insert(1, "Invoice", invoice)
                df_input.insert(2, "Fecha Factura", fecha_factura)
                df_merge = df_input.merge(df_tipo_servicio[['Tarifa', 'Tipo de Servicio']],on='Tarifa',how='left')

                df_input.insert(3, "Tipo de Servicio", df_merge["Tipo de Servicio"].fillna(""))

                start_row = next_row

                for row in dataframe_to_rows(df_input, index=False, header=False):
                    for col_idx, value in enumerate(row, 1):
                        ws.cell(row=next_row, column=col_idx, value=value)
                    next_row += 1

                for row in range(start_row, next_row):
                    for col_idx, style in enumerate(style_reference, 1):
                        cell = ws.cell(row=row, column=col_idx)
                        if style['fill']:
                            cell.fill = style['fill']
                        if style['font']:
                            cell.font = style['font']
                        if style['border']:
                            cell.border = style['border']
                        if style['alignment']:
                            cell.alignment = style['alignment']
                        if style['number_format']:
                            cell.number_format = style['number_format']

        wb.save(output_apm_excel)
        logging.info("Archivo base de Excel actualizado correctamente.")
        return True
        
    except Exception as e:
        logging.error(f"Error al procesar archivos Excel: {e}")
        return False

def apm_limpiar_input():
    """Limpia todos los archivos de la carpeta input"""
    try:
        for filename in os.listdir(input_folder):
            file_path = os.path.join(input_folder, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                logging.error(f'Error al eliminar {file_path}. Razón: {e}')
        logging.info("Carpeta input limpiada exitosamente")
        return True
    except Exception as e:
        logging.error(f"Error al limpiar la carpeta input: {e}")
        return False

def apm_descomprimir_zip(ruta_zip):
    with zipfile.ZipFile(ruta_zip, 'r') as zip_ref:
        zip_ref.extractall(input_folder)
    os.remove(ruta_zip)

def dpw_esperar_archivos(carpeta, tiempo_max_espera=6000, intervalo=10):
    """Espera hasta que haya archivos PDF y Excel en la carpeta"""
    tiempo_inicio = time.time()
    
    while (time.time() - tiempo_inicio) < tiempo_max_espera:
        archivos_pdf = [f for f in os.listdir(carpeta) if f.lower().endswith('.pdf')]
        archivos_excel = [f for f in os.listdir(carpeta) if f.lower().endswith('.xls')]
        
        if archivos_pdf and archivos_excel:
            return archivos_pdf, archivos_excel
        
        if int(time.time() - tiempo_inicio) % 30 == 0:
             logging.info(f"Esperando archivos... PDF: {len(archivos_pdf)}, Excel: {len(archivos_excel)}")
        time.sleep(intervalo)     
    
    logging.warning(f"Tiempo de espera agotado ({tiempo_max_espera} segundos). No se encontraron ambos archivos.")
    return None, None

def dpw_extraer_draft(nombre_archivo):
    """Extrae el código draft del nombre del archivo"""
    partes = nombre_archivo.split('_')
    if len(partes) > 1:
        return partes[1].split('.')[0]
    return None

# ==================== FUNCIONES PARA PDF DPW ====================
def dpw_extraer_texto_pdf(ruta_pdf):
    """Extrae todo el texto de un PDF"""
    try:
        with open(ruta_pdf, 'rb') as pdf_file:
            reader = PyPDF2.PdfReader(pdf_file)
            return "".join(page.extract_text() for page in reader.pages)
    except Exception as e:
        logging.error(f"Error al leer el PDF: {e}")
        return None

def dpw_extraer_cod_factura(texto_pdf):
    """Extrae el código de factura entre 'N° ' y 'FACTURA ELECTRONICA'"""
    patrones = [
        r"N°\s*([A-Za-z0-9-]+)\s*FACTURA ELECTRONICA",
        r"N°\s*([A-Za-z0-9-]+)\s*(?:FACTURA|BOLETA)"
    ]
    for patron in patrones:
        match = re.search(patron, texto_pdf)
        if match:
            return match.group(1)
    return None

def dpw_extraer_fecha(texto_pdf):
    """Extrae la fecha que sigue a 'R.U.C.'"""
    match = re.search(r"R\.U\.C\.\s*(\d{4}/\d{2}/\d{2})", texto_pdf)
    if match:
        try:
            return datetime.strptime(match.group(1), "%Y/%m/%d").strftime("%d-%m-%Y")
        except ValueError:
            logging.error(f"Formato de fecha no reconocido: {match.group(1)}")
    return None

def dpw_relacionar_archivos(archivos_pdf, archivos_excel, carpeta):
    """Relaciona archivos por su código draft"""
    relaciones = []
    
    # Procesar PDFs
    pdf_info = []
    for pdf in archivos_pdf:
        draft = dpw_extraer_draft(pdf)
        if draft:
            texto_pdf = dpw_extraer_texto_pdf(os.path.join(carpeta, pdf))
            if texto_pdf:
                pdf_info.append({
                    'nombre_original': pdf,
                    'draft': draft,
                    'cod_factura': dpw_extraer_cod_factura(texto_pdf),
                    'fecha': dpw_extraer_fecha(texto_pdf),
                    'texto_pdf': texto_pdf
                })
    
    # Procesar Excels
    excel_info = [{'nombre_original': excel, 'draft': dpw_extraer_draft(excel)} 
                 for excel in archivos_excel if dpw_extraer_draft(excel)]
    
    # Relacionar por draft
    for excel in excel_info:
        for pdf in pdf_info:
            if excel['draft'] == pdf['draft']:
                relaciones.append({'pdf': pdf, 'excel': excel})
                break
    
    return relaciones

def dpw_procesar_pdfs():
    """Procesa los archivos PDF y renombra los Excel"""
    os.makedirs(output_dpw_pdf, exist_ok=True)
    
    print("Esperando por archivos PDF y Excel...")

    # Esperar hasta que haya al menos un archivo PDF en la carpeta (máximo 30 segundos)
    tiempo_espera = 6000
    intervalo = 3
    tiempo_transcurrido = 0

    while tiempo_transcurrido < tiempo_espera:
        archivos = os.listdir(input_folder_dpw)
        hay_pdf = any(f.lower().endswith(".pdf") for f in archivos)
        hay_excel = any(f.lower().endswith((".xls", ".xlsx")) for f in archivos)
        if hay_pdf and hay_excel:
            break
        time.sleep(intervalo)
        tiempo_transcurrido += intervalo

    if not (hay_pdf and hay_excel):
        logging.warning("No se encontraron ambos archivos (PDF y Excel) en el tiempo límite.")
        return False
    
    logging.info("Archivos detectados. Iniciando procesamiento...")

    archivos_pdf, archivos_excel = dpw_esperar_archivos(input_folder_dpw)
    
    if not archivos_pdf or not archivos_excel:
        return False
    
    relaciones = dpw_relacionar_archivos(archivos_pdf, archivos_excel, input_folder_dpw)
    
    if not relaciones:
        logging.warning("No se pudo relacionar los archivos por código draft.")
        return False
    
    for relacion in relaciones:
        pdf_info = relacion['pdf']
        excel_info = relacion['excel']
        
        if not pdf_info['cod_factura'] or not pdf_info['fecha']:
            continue
        
        # Renombrar PDF
        nuevo_nombre_pdf = f"{pdf_info['cod_factura']}_{pdf_info['draft']}.pdf"
        ruta_pdf_original = os.path.join(input_folder_dpw, pdf_info['nombre_original'])
        ruta_pdf_nuevo = os.path.join(input_folder_dpw, nuevo_nombre_pdf)
        
        try:
            os.rename(ruta_pdf_original, ruta_pdf_nuevo)
            logging.info(f"PDF renombrado: {nuevo_nombre_pdf}")

        except Exception as e:
            logging.error(f"Error renombrando PDF: {e}")
            continue
        
        # Renombrar Excel
        nombre_base_excel = os.path.splitext(excel_info['nombre_original'])[0]
        nuevo_nombre_excel = f"{nombre_base_excel}_{pdf_info['cod_factura']}_{pdf_info['fecha']}.xls"
        ruta_excel_original = os.path.join(input_folder_dpw, excel_info['nombre_original'])
        ruta_excel_nuevo = os.path.join(input_folder_dpw, nuevo_nombre_excel)
        
        try:
            os.rename(ruta_excel_original, ruta_excel_nuevo)
            logging.info(f"Excel renombrado: {nuevo_nombre_excel}")
        except Exception as e:
            logging.error(f"Error renombrando Excel: {e}")
        
        # Mover PDF
        try:
            shutil.move(ruta_pdf_nuevo, os.path.join(output_dpw_pdf, nuevo_nombre_pdf))
            logging.info(f"PDF movido a: {output_dpw_pdf}")
        except Exception as e:
            logging.error(f"Error moviendo PDF: {e}")
    
    return True

# ==================== FUNCIONES PARA EXCEL DPW ====================
def dpw_procesar_excels():
    """Procesa archivos Excel manteniendo estilos correctamente"""
    try:
        # Cargar archivo base
        logging.info("Cargando archivo base...")
        wb = load_workbook(output_dpw_excel)
        
        if 'DPW' not in wb.sheetnames:
            raise ValueError("La hoja 'DPW' no existe en el archivo base.")
        
        ws = wb['DPW']
        
        # Encontrar última fila con datos
        logging.info("Buscando última fila con datos...")
        next_row = 2  # Asumiendo fila 1 es encabezado
        while ws.cell(row=next_row, column=1).value is not None:
            next_row += 1

        # Obtener estilos de referencia CORREGIDOS
        logging.info("Preparando estilos de referencia...")
        style_reference = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=2, column=col)
            style_reference.append({
                'fill': PatternFill(
                    start_color=cell.fill.start_color,
                    end_color=cell.fill.end_color,
                    fill_type=cell.fill.fill_type
                ) if cell.fill else PatternFill(),
                'font': Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    color=cell.font.color
                ) if cell.font else Font(),
                'border': Border(
                    left=Side(
                        border_style=cell.border.left.border_style,
                        color=cell.border.left.color
                    ) if cell.border.left else None,
                    right=Side(
                        border_style=cell.border.right.border_style,
                        color=cell.border.right.color
                    ) if cell.border.right else None,
                    top=Side(
                        border_style=cell.border.top.border_style,
                        color=cell.border.top.color
                    ) if cell.border.top else None,
                    bottom=Side(
                        border_style=cell.border.bottom.border_style,
                        color=cell.border.bottom.color
                    ) if cell.border.bottom else None
                ) if cell.border else Border(),
                'alignment': Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    wrap_text=cell.alignment.wrap_text
                ) if cell.alignment else Alignment(),
                'number_format': cell.number_format
            })

        # Procesar archivos
        archivos_procesados = 0
        for filename in os.listdir(input_folder_dpw):
            if filename.lower().endswith(".xls"):
                logging.info(f"\nProcesando archivo: {filename}")
                filepath = os.path.join(input_folder_dpw, filename)
                
                try:
                    # Leer archivo con engine adecuado
                    try:
                        df_input = pd.read_excel(filepath, header=2, engine='xlrd')
                    except:
                        df_input = pd.read_excel(filepath, header=2, engine='openpyxl')
                    
                    if df_input.empty:
                        logging.warning("  - Archivo vacío, saltando...")
                        continue

                    # Procesamiento de datos
                    df_input['BL'] = df_input.get('BL', '').astype(str).replace('nan', '').str.strip()

                    df_tipo_servicio = pd.read_excel(base_tipo_servicio_dpw, dtype=str)

                    df_input['Tariff'] = df_input['Tariff'].astype(str).str.strip()

                    df_tipo_servicio['Tariff'] = df_tipo_servicio['Tariff'].astype(str).str.strip()
                    
                    partes = filename.replace(".xls", "").split("_")
                    invoice = partes[-2]
                    draft = partes[-3]
                    fecha_factura = datetime.strptime(partes[-1], "%d-%m-%Y").strftime("%d/%m/%Y")

                    df_input['Amount'] = pd.to_numeric(df_input['Amount'], errors='coerce').fillna(0)

                    df_input['Description'] = df_input['Description'].astype(str).str.strip()

                    df_input = df_input[~(df_input['Tariff'].str.upper().str.contains("ZERO_STORAGE_SDT") | df_input['Description'].str.upper().isin(["TARIFA CERO / ZERO RATE", "TARIFA CERO / ZERO TARIFF"]) | (df_input['Amount'] == 0))]

                    df_input = df_input.reset_index(drop=True)

                    logging.info('Lógicas de cruces realizadas con éxito, insertando valores')
                    df_input.insert(0, "Proceso", df_input['BL'].apply(
                        lambda x: "EMBARQUE" if x == "" else "DESCARGA"))
                    df_input.insert(1, "Invoice", invoice)
                    df_input.insert(2, "Draft", draft)
                    df_input.insert(3, "Fecha Factura", fecha_factura)

                    df_merge = df_input.merge(df_tipo_servicio[['Tariff', 'Tipo de Servicio']],on='Tariff',how='left')

                    df_merge=df_merge.reset_index(drop=True)

                    df_input.insert(4, "Tipo de Servicio", df_merge["Tipo de Servicio"].fillna(""))

                    # Insertar datos con estilos
                    start_row = next_row
                    logging.info(f"  - Insertando {len(df_input)} registros desde fila {start_row}")
                    
                    for r_idx, row in enumerate(dataframe_to_rows(df_input, index=False, header=False), start=start_row):
                        for c_idx, value in enumerate(row, 1):
                            if c_idx - 1 < len(style_reference):
                                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                                style = style_reference[c_idx - 1]
                                
                                # Aplicar estilos de forma segura
                                cell.fill = style['fill']
                                cell.font = style['font']
                                cell.border = style['border']
                                cell.alignment = style['alignment']
                                cell.number_format = style['number_format']
                    
                    next_row += len(df_input)
                    archivos_procesados += 1
                    logging.info("  - Datos insertados correctamente")

                except Exception as e:
                    logging.error(f"  - Error procesando archivo: {str(e)}")
                    continue

        # Guardar cambios
        if archivos_procesados > 0:
            logging.info("\nGuardando cambios en el archivo base...")
            wb.save(output_dpw_excel)
            logging.info(f"  - Archivo base actualizado con {archivos_procesados} archivos procesados")
            return True
        else:
            logging.warning("\nNo se procesaron archivos para actualizar")
            return False

    except Exception as e:
        logging.error(f"\nError crítico: {str(e)}")
        return False

def dpw_limpiar_input():
    """Limpia todos los archivos de la carpeta input DPW"""
    try:
        for filename in os.listdir(input_folder_dpw):
            file_path = os.path.join(input_folder_dpw, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                logging.error(f'Error al eliminar {file_path}. Razón: {e}')
        logging.info("Carpeta input DPW limpiada exitosamente")
        return True
    except Exception as e:
        logging.error(f"Error al limpiar la carpeta input DPW: {e}")
        return False

def dpw_descomprimir_zip(ruta_zip):
    """Descomprime archivos ZIP para DPW"""
    with zipfile.ZipFile(ruta_zip, 'r') as zip_ref:
        zip_ref.extractall(input_folder_dpw)
    os.remove(ruta_zip)